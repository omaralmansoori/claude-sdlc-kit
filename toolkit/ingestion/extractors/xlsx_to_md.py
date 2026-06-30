#!/usr/bin/env python3
"""Convert an Excel workbook (.xlsx/.xls) to markdown — one table per sheet.

Why this exists: markitdown does not split a workbook per-sheet cleanly, and
spec corpora often ship requirement catalogs / data dictionaries / authority
matrices as multi-sheet workbooks where each sheet is its own entity. This
extractor renders **one markdown table per worksheet**, with the sheet name as
an H2 heading so the sheet name stays grep-able.

Behaviour:
  * Reads cell *values* only (``data_only=True``) — formulas are not evaluated
    here; the cached value Excel last stored is used.
  * Merged cells: the top-left value is repeated across the merged span so the
    table stays rectangular and readable.
  * Empty trailing rows/columns are trimmed.

Output goes to stdout by default (the ingest router captures it and prepends
the ``> Source:`` line), or to a file with ``-o``.

Requires ``openpyxl`` (see ../requirements.txt). The import is lazy so that
``--help`` works even when openpyxl is not installed.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path


def _escape(value: object) -> str:
    """Render a cell value as safe single-line markdown table text."""
    if value is None:
        return ""
    text = str(value)
    # Collapse newlines (markdown table cells are single-line) and escape pipes.
    text = text.replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
    text = text.replace("|", "\\|")
    return text.strip()


def _merged_value_map(worksheet) -> dict:
    """Map every covered cell coordinate -> the merged range's top-left value."""
    filled: dict[tuple[int, int], object] = {}
    for rng in worksheet.merged_cells.ranges:
        top_left = worksheet.cell(row=rng.min_row, column=rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                filled[(r, c)] = top_left
    return filled


def _sheet_to_markdown(worksheet) -> str:
    merged = _merged_value_map(worksheet)
    rows: list[list[str]] = []
    max_col = 0
    for row in worksheet.iter_rows():
        cells: list[str] = []
        for cell in row:
            value = merged.get((cell.row, cell.column), cell.value)
            cells.append(_escape(value))
        # Track the rightmost non-empty column to trim padding.
        last = 0
        for idx, text in enumerate(cells):
            if text:
                last = idx + 1
        max_col = max(max_col, last)
        rows.append(cells)

    # Trim trailing empty rows.
    while rows and not any(rows[-1]):
        rows.pop()

    if not rows or max_col == 0:
        return "_(empty sheet)_\n"

    # Normalise every row to max_col width.
    rows = [r[:max_col] + [""] * (max_col - len(r[:max_col])) for r in rows]

    header = rows[0]
    body = rows[1:]
    out: list[str] = []
    out.append("| " + " | ".join(header) + " |")
    out.append("| " + " | ".join(["---"] * max_col) + " |")
    for r in body:
        out.append("| " + " | ".join(r) + " |")
    return "\n".join(out) + "\n"


def workbook_to_markdown(path: Path) -> str:
    """Convert a workbook at ``path`` to a markdown string. Requires openpyxl."""
    try:
        import openpyxl  # noqa: PLC0415 — lazy so --help works without the dep
    except ImportError as exc:  # pragma: no cover - environment dependent
        raise SystemExit(
            "openpyxl is required for xlsx_to_md.py.\n"
            "Install it (and the rest of the pillar deps):\n"
            "    python3 -m pip install -r toolkit/ingestion/requirements.txt\n"
            "or run toolkit/ingestion/setup.sh"
        ) from exc

    # Not read_only: read-only worksheets do not expose merged_cells, which we
    # need to keep tables rectangular. data_only uses Excel's last cached value.
    workbook = openpyxl.load_workbook(filename=str(path), data_only=True)
    parts: list[str] = [f"# {path.stem}\n"]
    for name in workbook.sheetnames:
        worksheet = workbook[name]
        parts.append(f"## {name}\n")
        parts.append(_sheet_to_markdown(worksheet))
        parts.append("")  # blank line between sheets
    workbook.close()
    return "\n".join(parts).rstrip() + "\n"


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="xlsx_to_md.py",
        description="Convert an Excel workbook to markdown (one table per sheet).",
    )
    parser.add_argument("input", type=Path, help="Path to the .xlsx/.xls workbook.")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Write markdown here (default: stdout).",
    )
    args = parser.parse_args(argv)

    if not args.input.is_file():
        print(f"error: not a file: {args.input}", file=sys.stderr)
        return 2

    markdown = workbook_to_markdown(args.input)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(markdown, encoding="utf-8")
    else:
        sys.stdout.write(markdown)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
